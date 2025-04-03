####################################################################################################################################
# DELTA TEST SPECIFICATION
#
# WHAT: 
# - This tool is creating a delta, more exactly a report with differences found between two test specifications.
# - Tool is having two options, working with local files or directly with members directly from Integrity.
# - Be sure that your settings in Integrity at Host Name and Port are correct.
# 
# WHY: 
# - Easier way to see changes between implementations based on test specifications. 
# 
# AUTHOR:
# - Bogdan Vaslan/Falamas Ovidiu
# 
# INPUT:
# - When selected "Work with local files": two different files of test specification
# 
# OUTPUT:
# - A report with differences found between two test specifications
###################################################################################################################################

import tkinter as tk
from tkinter import *
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, colors
from openpyxl.styles.borders import Border, Side
from interface import *
import tempfile
import time
import os
import re

exp_string = tk.StringVar()

class Tab7():
    
    def __init__(self, tab):
        self.work_mode = ""
        self.modules_list = []
        self.project_list = []
        self.TS_ver = []
        self.flag_proj_received = False
        self.file_path = ""
        
        #color of background labels    
        fg_col = "#291b47"
        bg_col = "#fcfcfc"
        bt_col = "#f6f6f6"
        gold_color = "#f05e53"
     
        #set margin right (in all scripts use these lines)
        self.lable0 = tk.Label(tab, text = "      ", bg = bg_col)
        self.lable0.grid(row=0, column=4, padx=100)  

        # set margin left (in all scripts use these lines)
        self.lable1 = tk.Label(tab, text = "      ", bg = bg_col)
        self.lable1.grid(row=0, column=0, padx=80)
        
        #set margin right (in all scripts use these lines)
        self.lable2 = tk.Label(tab, text = "      ", bg = bg_col)
        self.lable2.grid(row=0, column=1)

        #set margin right (in all scripts use these lines)
        self.lable3 = tk.Label(tab, text = "      ", width="75", bg = bg_col)
        self.lable3.grid(row=0, column=2)

        self.lable4 = tk.Label(tab, cursor="hand2", text="Test Specification Delta", fg = fg_col, font = "Calibri 18 bold", bg = bg_col)
        self.lable4.bind ('<1>', lambda self: bt7_msgbox())
        self.lable4.grid(row=1, column=2)

        self.lable5 = tk.Label(tab, text = "      ", bg = bg_col)
        self.lable5.grid(row=2, column=0, pady=10)

        self.Bt_select_local = Button(tab, text="Work with local files", activebackground=gold_color, width=25, height=2, command=self.load_interface_local, bg = bt_col)
        self.Bt_select_local.grid(row=3, column=2, sticky=W)
        
        self.Bt_select_MKS = Button(tab, text="Work with MKS members", activebackground=gold_color, width=25, height=2, command=self.load_interface_MKS, bg = bt_col)
        self.Bt_select_MKS.grid(row=3, column=2, sticky=E)

        self.lable5 = tk.Label(tab, text = "      ", bg = bg_col)
        self.lable5.grid(row=4, column=0)

        self.lable6 = tk.Label(tab, text = "      ", bg = bg_col)
        self.lable6.grid(row=7, column=0)

        self.lable7 = tk.Label(tab, text = "      ", bg = bg_col)
        self.lable7.grid(row=10, column=0)

        self.lable8 = tk.Label(tab, text = "      ", bg = bg_col)
        self.lable8.grid(row=13, column=0)

        self.lable9 = tk.Label(tab, text = "      ", bg = bg_col)
        self.lable9.grid(row=15, column=0)

        self.exp_btn = Button(tab, text="Open report", activebackground=bt_col, width=25, height=2, command=self.open_exp, state="disabled", bg = bt_col) 
        self.btn_quit = Button(tab, text="Quit", activebackground=gold_color, width=25, height=2, command=tb_call_quit, bg = bt_col)        

        ##############################
        #MKS members area
        ##############################
        self.Bt_Projects = Button(tab, text="Load available projects", width=35, height=2, command=self.get_projects, bg = bt_col)
        
        self.selected_project = StringVar(tab)
        self.selected_project.trace("w", self.get_modules)
        self.selected_project.set("Select one...")
     

        self.selected_modules = StringVar(tab)
        self.selected_modules.trace("w", self.get_versions)
        self.selected_modules.set("Select one...")

        self.Ver_1 = StringVar(tab)
        self.Ver_1.set("Select one...")

        self.Ver_2 = StringVar(tab)
        self.Ver_2.set("Select one...")
   
        self.dropdown_projects = OptionMenu(tab, self.selected_project, self.project_list)
        self.dropdown_projects.configure(state="disable", width=35, pady=1, activebackground=gold_color, bg = bt_col, highlightbackground = bg_col)
        self.Lbl_projects = tk.Label(tab, text="Click to choose a desired project     ------------->", bg = bg_col,pady=7)
        
        self.dropdown_modules = OptionMenu(tab, self.selected_modules, self.modules_list)
        self.dropdown_modules.configure(state="disable", width=35, pady=1, activebackground=gold_color, bg = bt_col, highlightbackground = bg_col)
        self.Lbl_Module = tk.Label(tab, text="Click to select a specific module     ------------->", bg = bg_col,pady=7)
        
        self.select_ver_1 = OptionMenu(tab, self.Ver_1, self.TS_ver)
        self.select_ver_1.configure(state="disable", width=35, pady=1, activebackground=gold_color, bg = bt_col, highlightbackground = bg_col)
        self.Lbl_TS_ver_1 = tk.Label(tab, text="Select the 1st version of test spec.     ------------>", bg = bg_col,pady=7) 
        
        self.select_ver_2 = OptionMenu(tab, self.Ver_2, self.TS_ver)
        self.select_ver_2.configure(state="disable", width=35, pady=1, activebackground=gold_color, bg = bt_col, highlightbackground = bg_col)
        self.Lbl_TS_ver_2 = tk.Label(tab, text="Select the 2nd version of test spec.     ----------->", bg = bg_col,pady=7)
        
        self.Btn_generate_delta_MKS = Button(tab, text="Generate Delta", activebackground=bt_col, width=35, height=2, command=self.generate_delta_MKS, state="disabled", bg = bt_col)

        ##############################
        #Local files area
        ##############################
        self.local_Path1 = tk.StringVar(tab)
        self.local_Path1.set("")
        
        self.local_Path2 = tk.StringVar(tab)
        self.local_Path2.set("")
        
        self.Lbl_TS1 = tk.Label(tab, text = "1st TS file: ", bg = bg_col, pady=5)
        self.Ent_TS1 = Entry(tab, textvariable=self.local_Path1, width="65")
     
        self.Lbl_TS2 = tk.Label(tab, text = "2nd TS file: ", bg = bg_col, pady=10)
        self.Ent_TS2 = Entry(tab, textvariable=self.local_Path2, width="65")

        self.Btn_getPath1 = Button(tab, text="...", width=3, pady=-1, command=lambda: self.set_file(self.local_Path1), state="active", activebackground=bt_col, bg = bt_col)
        self.Btn_getPath2 = Button(tab, text="...", width=3, pady=-1, command=lambda: self.set_file2(self.local_Path2), state="active", activebackground=bt_col, bg = bt_col)
        
        self.Btn_generate_delta_local = Button(tab, text="Generate Delta", activebackground=bt_col, width=25, height=2, command=self.generate_delta_local, state="disabled", bg = bt_col)

    def load_interface_local(self):
        if self.work_mode == "":

            self.work_mode = "Local"

            # set current layout of LOCAL INTERFACE
            #addition of local interface
            self.Lbl_TS1.grid(row=5, column=2, sticky=W)
            self.Ent_TS1.grid(row=5, column=2)
            self.Btn_getPath1.grid(row=5, column=2, sticky=E)

            self.Lbl_TS2.grid(row=6, column=2, sticky=W)
            self.Ent_TS2.grid(row=6, column=2)
            self.Btn_getPath2.grid(row=6, column=2, sticky=E)

            self.Btn_generate_delta_local.grid(row=8, column=2)
            self.exp_btn.grid(row=11, column=2, sticky=W)
            self.btn_quit.grid(row=11, column=2, sticky=E)

            #keep the appearances from first page
            self.lable0.grid(row=0, column=3, padx=80)  
            self.lable1.grid(row=0, column=0, padx=80)
            self.lable2.grid(row=0, column=1)
            self.lable3.grid(row=0, column=2)
            self.lable4.grid(row=1, column=2)
            self.lable5.grid(row=2, column=0, pady=10)
            self.Bt_select_local.grid(row=3, column=2, sticky=W)
            self.Bt_select_MKS.grid(row=3, column=2, sticky=E)
            self.lable5.grid(row=4, column=0)
            self.lable6.grid(row=7, column=0)
            self.lable7.grid(row=10, column=0)
            self.lable8.grid(row=13, column=0)
            self.lable9.grid(row=15, column=0)

        elif self.work_mode == "MKS":
            
            self.work_mode = "Local"

            # hide previous layout of MKS INTERFACE
            self.Bt_Projects.grid_remove()
            self.Lbl_projects.grid_remove()
            self.dropdown_projects.grid_remove()
            self.Lbl_Module.grid_remove()
            self.dropdown_modules.grid_remove()  
            self.Lbl_TS_ver_1.grid_remove()
            self.select_ver_1.grid_remove()
            self.Lbl_TS_ver_2.grid_remove()
            self.select_ver_2.grid_remove()  
            self.Btn_generate_delta_MKS.grid_remove()
            self.exp_btn.grid_remove()
            self.btn_quit.grid_remove()           

            # set current layout of LOCAL INTERFACE
            #addition of local interface
            self.Lbl_TS1.grid(row=5, column=2, sticky=W)
            self.Ent_TS1.grid(row=5, column=2)
            self.Btn_getPath1.grid(row=5, column=2, sticky=E)

            self.Lbl_TS2.grid(row=6, column=2, sticky=W)
            self.Ent_TS2.grid(row=6, column=2)
            self.Btn_getPath2.grid(row=6, column=2, sticky=E)

            self.Btn_generate_delta_local.grid(row=8, column=2)
            self.exp_btn.grid(row=11, column=2, sticky=W)
            self.btn_quit.grid(row=11, column=2, sticky=E)

            #keep the appearances from first page
            self.lable0.grid(row=0, column=3, padx=80)  
            self.lable1.grid(row=0, column=0, padx=80)
            self.lable2.grid(row=0, column=1)
            self.lable3.grid(row=0, column=2)
            self.lable4.grid(row=1, column=2)
            self.lable5.grid(row=2, column=0, pady=10)
            self.Bt_select_local.grid(row=3, column=2, sticky=W)
            self.Bt_select_MKS.grid(row=3, column=2, sticky=E)
            self.lable5.grid(row=4, column=0)
            self.lable6.grid(row=7, column=0)
            self.lable7.grid(row=10, column=0)
            self.lable8.grid(row=13, column=0)
            self.lable9.grid(row=15, column=0)

    def load_interface_MKS(self):
        if self.work_mode == "":

            self.work_mode = "MKS"
            
            # set current layout of MKS INTERFACE
            #addition of mks interface
            self.Bt_Projects.grid(row=5, column=2)
    
            self.Lbl_projects.grid(row=8, column=2, sticky=W)
            self.dropdown_projects.grid(row=8, column=2, sticky=E)
            
            self.Lbl_Module.grid(row=9, column=2, sticky=W)
            self.dropdown_modules.grid(row=9, column=2, sticky=E)    
            
            self.Lbl_TS_ver_1.grid(row=10, column=2, sticky=W)
            self.select_ver_1.grid(row=10, column=2, sticky=E) 
            
            self.Lbl_TS_ver_2.grid(row=11, column=2, sticky=W)
            self.select_ver_2.grid(row=11, column=2, sticky=E)  
            
            self.Btn_generate_delta_MKS.grid(row=14, column=2)
            self.exp_btn.grid(row=16, column=2, sticky=W)
            self.btn_quit.grid(row=16, column=2, sticky=E)

            #keep the appearances from first page
            self.lable0.grid(row=0, column=3, padx=80)  
            self.lable1.grid(row=0, column=0, padx=80)
            self.lable2.grid(row=0, column=1)
            self.lable3.grid(row=0, column=2)
            self.lable4.grid(row=1, column=2)
            self.lable5.grid(row=2, column=0, pady=10)
            self.Bt_select_local.grid(row=3, column=2, sticky=W)
            self.Bt_select_MKS.grid(row=3, column=2, sticky=E)
            self.lable5.grid(row=4, column=0)
            self.lable6.grid(row=7, column=0)
            self.lable7.grid(row=10, column=0)
            self.lable8.grid(row=13, column=0)
            self.lable9.grid(row=15, column=0)

        elif self.work_mode == "Local":

            self.work_mode = "MKS"

            #hide previous layout of LOCAL INTERFACE
            self.Lbl_TS1.grid_remove()
            self.Ent_TS1.grid_remove()
            self.Btn_getPath1.grid_remove()
            self.Lbl_TS2.grid_remove()
            self.Ent_TS2.grid_remove()
            self.Btn_getPath2.grid_remove()
            self.Btn_generate_delta_local.grid_remove()
            self.exp_btn.grid_remove()
            self.btn_quit.grid_remove()  

            # set current layout of MKS INTERFACE
            #addition of mks interface
            self.Bt_Projects.grid(row=5, column=2)
    
            self.Lbl_projects.grid(row=8, column=2, sticky=W)
            self.dropdown_projects.grid(row=8, column=2, sticky=E)
            
            self.Lbl_Module.grid(row=9, column=2, sticky=W)
            self.dropdown_modules.grid(row=9, column=2, sticky=E)    
            
            self.Lbl_TS_ver_1.grid(row=10, column=2, sticky=W)
            self.select_ver_1.grid(row=10, column=2, sticky=E) 
            
            self.Lbl_TS_ver_2.grid(row=11, column=2, sticky=W)
            self.select_ver_2.grid(row=11, column=2, sticky=E)  
            
            self.Btn_generate_delta_MKS.grid(row=14, column=2)
            self.exp_btn.grid(row=16, column=2, sticky=W)
            self.btn_quit.grid(row=16, column=2, sticky=E)

            #keep the appearances from first page
            self.lable0.grid(row=0, column=3, padx=80)  
            self.lable1.grid(row=0, column=0, padx=80)
            self.lable2.grid(row=0, column=1)
            self.lable3.grid(row=0, column=2)
            self.lable4.grid(row=1, column=2)
            self.lable5.grid(row=2, column=0, pady=10)
            self.Bt_select_local.grid(row=3, column=2, sticky=W)
            self.Bt_select_MKS.grid(row=3, column=2, sticky=E)
            self.lable5.grid(row=4, column=0)
            self.lable6.grid(row=7, column=0)
            self.lable7.grid(row=10, column=0)
            self.lable8.grid(row=13, column=0)
            self.lable9.grid(row=15, column=0)

    def set_file(self, output):
        try:
            path = tk.filedialog.askopenfilename(title = "Select TS export file", filetypes=[("Excel","*.xlsm")])
            output.set(path)
        except:
            messagebox.showerror("Unexpected error:", sys.exc_info()[0])

    def set_file2(self, output):
        try:
            path = tk.filedialog.askopenfilename(title = "Select TS export file", filetypes=[("Excel","*.xlsm")])
            output.set(path)
            self.Btn_generate_delta_local["state"]="active"
        except:
            messagebox.showerror("Unexpected error:", sys.exc_info()[0])        

    def MKS_get_project (self):
        stream = os.popen("si projects")
        output = stream.read()
        lista = output.split()
        
        for el, project in enumerate(lista):
            if "SWC_" in project:
                lista[el] = project.replace ("/project.pj", "")[1:]
            else:
                del lista[el]
        return lista
           
    def MKS_get_modules(self):
        stream = os.popen("si viewproject --project=#/" + self.selected_project.get() + "#01_PROD/40_T --no")
        output = stream.read().split()
        lista = output[::2]
        modul_list = []
        for el, modul in enumerate(lista):
            if "T_" in modul:
                modul_list.append(modul.replace("/project.pj", "")[2:])
        return modul_list
      
    def MKS_get_versions (self):
        new_output = []
        stream = os.popen("si viewproject --project=#/" + self.selected_project.get() + "#01_PROD/40_T/T_" + self.selected_modules.get() + "/40_T/10_TS")
        tmp_output = stream.read()
        if "TS_" not in tmp_output:
            return new_output
        output = tmp_output.split()
        
        
        stream = os.popen("si viewhistory --project=#/" + self.selected_project.get() + "#01_PROD/40_T/T_" + self.selected_modules.get() + "/40_T/10_TS --fields=revision " + output[0])
        new_output = [x for x in stream.read().split("\n")[1:-2]]
        return new_output
            
    def generate_delta_MKS(self):   
        if self.Ver_1.get() == 'Select TS Version' or self.Ver_2.get() == 'Select TS Version':
            messagebox.showinfo("Warning","Select the TS export versions")
        elif self.Ver_1.get() == self.Ver_2.get():
            messagebox.showinfo("Warning","The TS versions can not match")
        else:
            tmp_folder = self.get_save_path("Select path for saving Test Specification files")            
            if tmp_folder:

                export_MKS_1 = os.popen("si projectco --nolock --project=#/" + self.selected_project.get() + "#01_PROD/40_T/T_" + self.selected_modules.get() + "/40_T/10_TS --targetFile='" + tmp_folder + r"/testspec_1.xlsm' --revision=" + self.Ver_1.get() + " TS_SWC_" + self.selected_modules.get() + ".xlsm")
                export_MKS_2 = os.popen("si projectco --nolock --project=#/" + self.selected_project.get() + "#01_PROD/40_T/T_" + self.selected_modules.get() + "/40_T/10_TS --targetFile='" + tmp_folder + r"/testspec_2.xlsm' --revision=" + self.Ver_2.get() + " TS_SWC_" + self.selected_modules.get() + ".xlsm")
            
                time.sleep(5)
            
                self.process_files(tmp_folder + r"/testspec_1.xlsm", tmp_folder + r"/testspec_2.xlsm")
               
    def generate_delta_local(self):
        if self.local_Path1.get() == "" or self.local_Path2.get() == "":
            messagebox.showinfo("Information", "One of the path is empty")
            return
        if self.local_Path1.get() == self.local_Path2.get():
            messagebox.showinfo("Information", "Please select two different files")

        #generate overview    
        self.process_files(self.local_Path1.get(), self.local_Path2.get())
               
    def get_projects(self):
        if self.flag_proj_received == False:
            
            self.project_list = self.MKS_get_project()
            
            self.selected_project.set('Select one...')
            self.dropdown_projects['menu'].delete(0, 'end')
            self.dropdown_projects.configure(state="active")
    
            for choice in self.project_list:
                 self.dropdown_projects['menu'].add_command(label=choice, command=tk._setit(self.selected_project, choice))
            self.flag_proj_received = True
            
    def get_modules(self, *args):
        
        if self.selected_project.get() != "Select one...":
            self.modules_list = self.MKS_get_modules()
            
            if len(self.modules_list) > 0:
                self.selected_modules.set('Select one...')
                self.dropdown_modules['menu'].delete(0, 'end')
                self.dropdown_modules.configure(state="active")
                
                self.select_ver_1.configure(state="disabled")
                self.select_ver_2.configure(state="disabled")
                self.Ver_1.set('Select one...')
                self.Ver_2.set('Select one...')
                
                self.Btn_generate_delta_MKS["state"]="disabled"

                for choice in self.modules_list:
                    self.dropdown_modules['menu'].add_command(label=choice, command=tk._setit(self.selected_modules, choice))
            else:

                messagebox.showinfo("Warning","The selected project does not have regular modules!")

                self.selected_modules.set('Select one...')
                self.dropdown_modules.configure(state="disabled")

                self.select_ver_1.configure(state="disabled")
                self.select_ver_2.configure(state="disabled")
                self.Ver_1.set('Select one...')
                self.Ver_2.set('Select one...')
    
    def get_versions(self, *args):
        if self.selected_modules.get() != "Select one...":
            self.TS_ver = self.MKS_get_versions()
            
            self.Ver_1.set('Select one...')
            self.Ver_2.set('Select one...')
            self.select_ver_1['menu'].delete(0, 'end')
            self.select_ver_2['menu'].delete(0, 'end')
            
            if len(self.TS_ver) > 1:
                self.select_ver_1.configure(state="active")
                self.select_ver_2.configure(state="active")
                self.Btn_generate_delta_MKS["state"]="active"
            elif len(self.TS_ver) == 1:
                self.Btn_generate_delta_MKS["state"]="disabled"
                messagebox.showinfo("Warning","The selected module contains only one test specification export!")
            else:
                self.Btn_generate_delta_MKS["state"]="disabled"
                messagebox.showinfo("Warning","The selected module contains no test specification export!")
                self.Ver_1.set('No TS available')   
                self.Ver_2.set('No TS available')
            

            for choice in self.TS_ver:
                 self.select_ver_1['menu'].add_command(label=choice, command=tk._setit(self.Ver_1, choice))
                 self.select_ver_2['menu'].add_command(label=choice, command=tk._setit(self.Ver_2, choice))
    
    def read_file(self, file_path):
        wb = load_workbook(file_path)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        TS_version = ws.cell(13,13).value
        project_name  = ws.cell(38,1).value
        project_name = project_name[project_name.index("TS_")+3:]

        ws = wb[sheets[1]]
        
        dimensions = ws.dimensions
        number_rows=int(dimensions.split(":")[1][1:])
        #########
        # translate header
        #########
        dictionry_header = {"Object Heading": "Test case name", 
                            "Object Text": "Description",
                            "Linked requirements": "Linked requirements",
                            "Test Method":"Test Method",
                            "Preconditions":"Preconditions",
                            "Actions":"Actions",
                            "Expected Results":"Expected results",
                            "Test Specification":"Actions",
                            "Pass Conditions":"Expected results"}
        heading_list = []
        for i in range(3,15):
            if ws.cell(2,i).value != None:
                heading_list.append(dictionry_header.get(ws.cell(2,i).value,"Not available"))
        
        #check if the TS export contains the new attributes. 
        #Unfortunately the last row is not correct and I have to use this hack
        
        offset = 0
        check_new_template = ws.cell(2,9).value
        if check_new_template != None:
            offset = 1
        
        content = {}
        for i in range(3,number_rows):
            test_case_id = int(ws.cell(i,1).value)
            if ws.cell(i,2).value == "Test case":
                content[test_case_id] =  dict()
                for j in range(len(heading_list)):
                    content[test_case_id][heading_list[j]] = ws.cell(i,j+3).value
        return TS_version, content, project_name
       
    def get_save_path(self, message):
        file_path = filedialog.askdirectory(title = message)
        
        if file_path == "":
            return None
        else:
            return file_path
        
    def process_files(self, file1, file2):
        link_ft = Font(underline= "single", color = colors.BLUE)
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        
        new_version, new_content, project_new = self.read_file(file1)
        old_version, old_content, project_old = self.read_file(file2)
        
        if int(new_version.split(".")[1]) < int(old_version.split(".")[1]):
            new_version, old_version = (old_version , new_version)
            new_content, old_content = (old_content, new_content)
        
        new_content_keys = new_content.keys()
        old_content_keys = old_content.keys()
        
        deleted_TCs = old_content_keys-new_content_keys
        new_TCs = new_content_keys-old_content_keys
        same_TCs = new_content_keys-new_TCs
        changed_TCs = list()
        
        for TC in same_TCs:
            for key in new_content[TC].keys():
                if (key in new_content[TC].keys() and key in old_content[TC].keys()) and new_content[TC][key] != old_content[TC][key] and key != "Not available":
                    changed_TCs.append(TC)
                    break
        
        resulted_wb = Workbook()
        resulted_wb.active.title = "Overview Report"
        
        resulted_wb.active["D4"] = "Delta report for " + project_new        
        resulted_wb.active["D4"].font = Font(size="20", bold=True)
        resulted_wb.active["C5"] = "based on TS exported from TPT versions: {0} vs {1}".format(new_version, old_version)
        resulted_wb.active["C5"].font = Font(size="20", bold=False)
        
        type_sheets = ["Deleted TCS", "New TCs", "Changed TCs"]
        list_scenarios = [deleted_TCs, new_TCs, changed_TCs]
        start_row = 10
        start_column = 4
        
        for pos, item in enumerate(list_scenarios):
            resulted_wb["Overview Report"].cell(start_row, start_column).value = type_sheets[pos]
            resulted_wb["Overview Report"].cell(start_row, start_column).alignment = Alignment(horizontal="center", vertical="center")
            resulted_wb["Overview Report"].cell(start_row, start_column).border = thin_border
            resulted_wb["Overview Report"].column_dimensions["D"].width = "20"
            if len(item) <1: 
                resulted_wb["Overview Report"].cell(start_row, start_column+1).value = "-"
                resulted_wb["Overview Report"].cell(start_row, start_column+1).alignment = Alignment(horizontal="center")
                resulted_wb["Overview Report"].cell(start_row, start_column+1).border = thin_border
                start_row += 1
            else:
                resulted_wb["Overview Report"].merge_cells(None, start_row, start_column, start_row+len(item)-1, start_column)
                
                    
                resulted_wb.create_sheet(type_sheets[pos])
                current_ws = resulted_wb[type_sheets[pos]]
                current_ws.column_dimensions["B"].width = "20"
                current_ws.column_dimensions["C"].width = "20"
                current_ws.column_dimensions["D"].width = "70"
                current_ws.column_dimensions["E"].width = "70"
                for pos_n,TC in enumerate(item):
                    resulted_wb["Overview Report"].cell(start_row, start_column+1).value = TC
                    resulted_wb["Overview Report"].cell(start_row, start_column+1).alignment = Alignment(horizontal="center")
                    resulted_wb["Overview Report"].cell(start_row, start_column+1).border = thin_border
                    
                    
                    key_index = 0
                    current_ws.cell(3+pos_n*10+key_index, 2).value = "Test case ID= " + str(TC)
                    current_ws.cell(3+pos_n*10+key_index, 2).border = thin_border
                    
                    link = "#'"+ type_sheets[pos] + "'!B" + str(3+pos_n*10+key_index)
        
                    resulted_wb["Overview Report"].cell(start_row, start_column+1).hyperlink = (link) 
                    resulted_wb["Overview Report"].cell(start_row, start_column+1).font = link_ft
                    start_row += 1
                         
                    if pos == 0:
                        container = old_content
                    else:
                        container = new_content
                    if pos == 2:
                        current_ws.cell(3+pos_n*10+key_index-1, 4).value = "New content"
                        current_ws.cell(3+pos_n*10+key_index-1, 4).border = thin_border
                        current_ws.cell(3+pos_n*10+key_index-1, 5).value = "Old content"
                        current_ws.cell(3+pos_n*10+key_index-1, 5).border = thin_border
                    for key in container[TC].keys():
                        if key != "Not available":
                            current_ws.cell(3+pos_n*10+key_index, 3).value = key
                            current_ws.cell(3+pos_n*10+key_index, 3).border = thin_border
                            current_ws.cell(3+pos_n*10+key_index, 3).alignment = Alignment(wrap_text=True,vertical='center') 
                            current_ws.cell(3+pos_n*10+key_index, 4).value = container[TC][key]
                            current_ws.cell(3+pos_n*10+key_index, 4).border = thin_border
                            current_ws.cell(3+pos_n*10+key_index, 4).alignment = Alignment(wrap_text=True,vertical='top') 
                            if pos ==2:
                                current_ws.cell(3+pos_n*10+key_index, 5).value = old_content[TC].get(key,"-")
                                current_ws.cell(3+pos_n*10+key_index, 5).border = thin_border
                                current_ws.cell(3+pos_n*10+key_index, 5).alignment = Alignment(wrap_text=True,vertical='top') 
                                if (key in new_content[TC].keys() and key in old_content[TC].keys()) and new_content[TC][key] != old_content[TC][key]:
                                    current_ws.cell(3+pos_n*10+key_index, 4).font = Font(color="00b050") 
                                    current_ws.cell(3+pos_n*10+key_index, 5).font = Font(color="FF0000")
                            key_index +=1
                        
                    current_ws.sheet_view.zoomScale = 85
                    current_ws.column_dimensions["A"].width = "20"
                    current_ws.cell(1,1).value = "Back to first page"
                    current_ws.cell(1,1).hyperlink = "#'Overview Report'!A1"
                    current_ws.cell(1,1).font = link_ft

        self.file_path = file2.partition('/testspec')[0]    

        if os.path.isdir(self.file_path): 
            resulted_wb.save(self.file_path + r"/Overview_report.xlsx")
            messagebox.showinfo("Information", "Done! Your overview report is ready!")
            exp_string.set(self.file_path + '/Overview_report.xlsx')
            self.exp_btn["state"]="active"            


    def open_exp(self):

        excel_path = exp_string.get()

        if not excel_path:
            messagebox.showerror("FileNotFoundError","No excel file was selected")
        else:
            os.startfile(excel_path)